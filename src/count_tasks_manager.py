"""
这个文件用来

储存用户新建的count_tasks

"""

from openpyxl import load_workbook
from services import config
import os
from openpyxl import Workbook
from services import TaskTimer, TaskRecorder, TimerStatus
import user_json_path


class CountTasksManager:
    def __init__(self):
        self.tasks = config.get("count_tasks")()
        # 指定用于存储 count_tasks 数据的 Excel 文件路径
        self.excel_path = "data/count_tasks_data.xlsx"
        self.load_completion_counts()

    def delete_task_records(self, task_id):
        excel_path = "data/count_tasks_data.xlsx"  # Excel 文件路径
        try:
            wb = load_workbook(excel_path)
            ws = wb.active
            for row in range(ws.max_row, 1, -1):
                if (
                    ws.cell(row=row, column=1).value == task_id
                ):  # 假设任务ID在第一列
                    ws.delete_rows(row)
            wb.save(excel_path)
        except Exception as e:
            print(f"Error deleting task records from Excel: {e}")

    def load_completion_counts(self):
        # 加载 Excel 文件中的完成次数
        if not os.path.exists(self.excel_path):
            # 如果文件不存在，创建一个空的 Excel 文件
            wb = Workbook()
            ws = wb.active
            ws.append(["Task ID", "Completion Count"])  # 添加表头
            wb.save(self.excel_path)
            return
        wb = load_workbook(self.excel_path)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            task_id, completion_count = row
            for task in self.tasks:
                if task["id"] == task_id:
                    task["completion_count"] = completion_count
                    break

    def save_completion_count(self, task_id):
        # 保存完成次数到 Excel 文件
        wb = (
            load_workbook(self.excel_path)
            if os.path.exists(self.excel_path)
            else Workbook()
        )
        ws = wb.active
        found = False
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=2, max_col=2, values_only=True),
            start=2,
        ):  # 从2开始枚举
            if row[0] == task_id:
                ws.cell(
                    row=row_idx, column=2, value=row[1] + 1
                )  # 使用正确的行号
                found = True
                break
        if not found:
            ws.append([task_id, 1])
        wb.save(self.excel_path)

    def increment_completion_count(self, task_id):
        # 增加任务的完成次数
        for task in self.tasks:
            if task["id"] == task_id:
                task["completion_count"] = (
                    task.get("completion_count", 0) + 1
                )
                self.save_completion_count(task_id)
                break
