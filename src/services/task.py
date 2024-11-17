"""_summary_
    任务计时器
"""
from .utils.helper import *
from .config import *
from openpyxl import Workbook,load_workbook
import os
from threading import Thread,Event
from time import sleep

__all__= ('TimerStatus','TaskTimer','TaskRecord','TaskRecorder')

TimerStatus = namedtuple('TimerStatus',('task_id','countdown','start','end','running','used_time')) # used_time指有效时间 暂停时不算
class TaskTimer:
    instances    = []
    exit:Event   = Event()
    thread:Thread
    def __init__(self,task_id:int,countdown_time:Optional[timedelta]=None):
        self.task_id = task_id # 任务id
        self.countdown = countdown_time  # 倒计时设定时间,None表示正计时
        self.start = datetime.now() # 任务开始时间
        self.running = True   # 任务是暂停还是运行
        self.end = None  # 任务结束时间,None表示未结束
        self._last = [datetime.now(),timedelta(0)] # 上次暂停/恢复时的 时间和已用时间
        TaskTimer.instances.append(self)
    
    def __del__(self):
        if self in TaskTimer.instances:
            TaskTimer.instances.remove(self)
    
    def resume(self)->None:
        if self.end is not None:
            return
        if self.running:
            return
        self.running = True
        self._last[0] = datetime.now()
    
    def pause(self)->None:
        if self.end is not None:
            return
        if not self.running:
            return
        self.running = False
        self._last[1] += datetime.now() - self._last[0]
        self._last[0] = datetime.now()
    
    def stop(self)->None:
        if self.end is not None:
            return
        self.pause()
        self.end = datetime.now()
        
    @property
    def _used(self)->timedelta: # 理论已用时间，不会改变计时器状态
        return self._last[1] + datetime.now() - self._last[0] if self.running else self._last[1]
    
    @property
    def status(self)->TimerStatus: # 更新并获取状态
        if (self.end is None) and (self.countdown is not None) and (self._used > self.countdown): # 未结束+倒计时+超时
            self.stop() # 自动结束
        return TimerStatus(self.task_id,self.countdown,self.start,self.end,self.running,self._used)
    
    def __str__(self)->str:
        return f"task_id: {self.status.task_id} -- {self.status.used_time} / {self.status.countdown}"
    
    def __repr__(self)->str:
        status=self.status # 注意这里会update
        return f'任务id:{status.task_id} 倒计时:{status.countdown}\n开始时间:{status.start} 结束时间:{status.end}\n是否运行:{status.running} 已用时间:{status.used_time}'
    
    @staticmethod
    def _run(): # 类的子线程使用这个方法
        def wq(i:int): # 写入表格并删除实例
            TaskRecorder.add_records(TaskRecorder.rec_split(TaskRecorder.timer_record(TaskTimer.instances[i].status)))
            # print(f"[reverse]recorded[/]: {TaskTimer.instances[i].status.task_id}")
            del TaskTimer.instances[i]
        
        interval=config.get('settings','advanced','timer_interval_sec')()
        while not TaskTimer.exit.is_set():
            for i in range(len(TaskTimer.instances) - 1, -1, -1): #遍历类的所有实例
                if TaskTimer.instances[i].status.end is not None:
                    wq(i)
                else:
                    pass
                    # print(TaskTimer.instances[i].__str__())
            # sleep(interval) #TODO 
        
        for i in range(len(TaskTimer.instances) - 1, -1, -1): # exit时停止并保存所有计时器
            TaskTimer.instances[i].stop()
            wq(i)
        
        assert TaskTimer.instances == [],"[reverse magenta]exit with timers unrecorded[/]"
        print("[reverse]safe exit[/]")

    
    @staticmethod
    def run_thread(): # 主线程使用这个方法
        TaskTimer.thread=Thread(target=TaskTimer._run)
        TaskTimer.exit.clear()
        TaskTimer.thread.start()
    
    @staticmethod
    def stop_thread(): # 主线程使用这个方法
        TaskTimer.exit.set()
        TaskTimer.thread.join()
        print("[reverse]join[/]")


# if __name__ == '__main__':
#     t=TaskTimer(1)
#     cdt=TaskTimer(2,countdown_time=timedelta(seconds=5))
#     breakpoint()
    
    
    
"""_summary_
    任务记录表交互
"""


TaskRecord=namedtuple('TaskRecord',('task_id','start','end','valid_time','note'))
class TaskRecorder:
    @staticmethod
    def floor_dtm(time:datetime)->datetime: # 削去微秒
        return time.replace(microsecond=0)
    
    @staticmethod
    def floor_dur(time:timedelta)->timedelta: # 削去微秒
        return timedelta(seconds=time.total_seconds()//1)
    
    @staticmethod
    def timer_record(status:TimerStatus,note:str=None)->TaskRecord: # TimerStatus 2 TaskRecord
        assert status.running is False, 'Timer not paused'
        assert status.end is not None, 'Timer not ended'
        
        start=TaskRecorder.floor_dtm(status.start)
        end=TaskRecorder.floor_dtm(status.end)
        valid_time=TaskRecorder.floor_dur(status.used_time)
        return TaskRecord(status.task_id,start,end,valid_time,note)
    
    @staticmethod
    def counter_record(task_id:int,time:datetime,note:str=None)->TaskRecord:
        tm=TaskRecorder.floor_dtm(time)
        return TaskRecord(task_id,tm,tm,None,note)
    
    @staticmethod
    def rec_split(rec:TaskRecord)->list[TaskRecord]: # 如果TaskRecord跨天，则拆分成多个在同一天的记录
        if rec.valid_time is None: # 计次任务
            return [rec]
        if rec.start.date()==rec.end.date(): # 同一天
            return [rec]
        assert rec.start.date()<rec.end.date(), 'TaskRecord not cross day'
        # 跨天
        start:datetime=rec.start
        end:datetime=rec.end
        elapsed:timedelta = end - start
        
        left:datetime = start
        right:datetime = start.replace(hour=23,minute=59,second=59,microsecond=999999)
        ret=[]
        while True:
            # 添加元素
            weight:float = (right-left)/elapsed
            valid_time:timedelta = TaskRecorder.floor_dur(rec.valid_time*weight)
            record:TaskRecord = TaskRecord(rec.task_id,TaskRecorder.floor_dtm(left),TaskRecorder.floor_dtm(right),valid_time,rec.note)
            ret.append(record)
            # 更改 left right
            left:datetime = right.replace(hour=0,minute=0,second=0,microsecond=0) + timedelta(days=1)
            if left.date()>end.date():
                break
            if left.date()==end.date():
                right=end
            else:
                right=left.replace(hour=23,minute=59,second=59,microsecond=999999)
        return ret
    
    # excel存入
    path='data/record.xlsx'
    
    @staticmethod
    def add_record(record:TaskRecord):
        assert record.start.date()==record.end.date(), 'TaskRecord not splitted'
        date=record.start.date()
        start=record.start.time()
        end=record.end.time()
        date=record.start.date()
        sheet_name='taskRecords'
    
        wb = load_workbook(TaskRecorder.path) if os.path.exists(TaskRecorder.path) else Workbook()
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
        
        ws.append([record.task_id,
                   date,
                   start,
                   end,
                   record.valid_time,
                   record.note
                   ])
        
        wb.save(TaskRecorder.path)
    
    @staticmethod
    def add_records(records:list[TaskRecord]):
        for rec in records:
            TaskRecorder.add_record(rec)

if __name__ == '__main__':
    
    import random
    VARY = 5e4
    for _ in range(10):
        dur=timedelta(seconds=VARY*random.random())
        shift=timedelta(seconds=200*VARY*(random.random()-.5))
        start=datetime.now()-0.7*dur+shift
        end = datetime.now()+0.7*dur+shift
        status=TimerStatus(1,2*dur,start,end,False,dur)
        record=TaskRecorder.timer_record(status)
        records=TaskRecorder.rec_split(record)
        print(records)
        TaskRecorder.add_records(records)
    
    breakpoint()